# -*- coding: utf-8 -*-
"""
模型排行榜聚合抓取脚本
功能: 抓取多个 AI 模型排行榜数据到 data/ 目录，供查询脚本使用
数据源:
  1. LMArena (arena.ai)        - HF 官方数据集 parquet (Elo/rating + rank, 29 分类)
  2. LiveBench                 - 官方公开 CSV + categories JSON
  3. SWE-bench                 - 官网 HTML 内嵌 JSON (代码能力)
  4. EQ-Bench                  - 官方 JS 数据文件 (情感智能 Elo)
  5. Artificial Analysis       - 官方 API (需 key) 或 首页 HTML 内嵌 JSON (Intelligence Index)
  6. OpenCompass 司南          - 上海AI实验室 OSS 公开 JSON (7 能力表)
  7. SuperCLUE                - 官方公开 XLSX (总榜/推理/开源榜)
  8. MMLU-Pro                 - HF Space gradio config (262 模型 × 14 学科)
  9. GPQA Diamond             - evals.report 官方验证榜 (84 模型)
  10. Terminal-Bench 2.1      - evals.report 官方验证榜 (CLI 智能体)
  11. SWE-bench Pro           - evals.report 官方验证榜 (企业级代码)
  12. BrowseComp-Plus         - HF Space 官方 leaderboard.json (网页浏览)
用法:
  python fetch_all.py               # 抓取全部榜单
  python fetch_all.py lmarena       # 只抓某个榜单
  python fetch_all.py --force       # 强制刷新 (默认有缓存则跳过)
"""
import os
import re
import sys
import json
import ssl
import gzip
import zlib
import time
import urllib.request
import urllib.error

# 强制 UTF-8 输出，避免 Windows GBK 控制台编码错误
try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

# ============ 配置 ============
SKILL_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(SKILL_DIR, 'data')
AUTH_FILE = os.path.join(SKILL_DIR, 'model_leaderboards_auth.json')
os.makedirs(DATA_DIR, exist_ok=True)

UA = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36'
CTX = ssl.create_default_context()
CTX.check_hostname = False
CTX.verify_mode = ssl.CERT_NONE

# LiveBench 已知 release 版本（页面 UI 按钮，日期升序）
LIVEBENCH_RELEASES = [
    '2024_06_24', '2024_07_26', '2024_08_31', '2024_11_25',
    '2025_04_02', '2025_04_25', '2025_05_30', '2025_11_25',
    '2025_12_23', '2026_01_08', '2026_06_25',
]


# ============ 工具函数 ============

def _fetch(url, timeout=25, retries=3, headers_extra=None):
    """抓取 URL，自动解压 gzip/deflate，失败重试"""
    headers = {
        'User-Agent': UA,
        'Accept': '*/*',
        'Accept-Encoding': 'gzip, deflate',
        'Accept-Language': 'en-US,en;q=0.9',
    }
    if headers_extra:
        headers.update(headers_extra)
    for attempt in range(1, retries + 1):
        try:
            req = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, context=CTX, timeout=timeout) as resp:
                raw = resp.read()
                enc = resp.headers.get('Content-Encoding', '')
                if enc == 'gzip':
                    raw = gzip.decompress(raw)
                elif enc == 'deflate':
                    try:
                        raw = zlib.decompress(raw)
                    except zlib.error:
                        raw = zlib.decompress(raw, -zlib.MAX_WBITS)
                return raw
        except urllib.error.HTTPError as e:
            if e.code in (429, 500, 502, 503, 504) and attempt < retries:
                time.sleep(2 ** attempt)
                continue
            print(f'    [ERROR] HTTP {e.code}: {url}')
            return None
        except Exception as e:
            if attempt < retries:
                time.sleep(2 ** attempt)
                continue
            print(f'    [ERROR] {type(e).__name__}: {e} -> {url}')
            return None
    return None


def _save(name, data):
    path = os.path.join(DATA_DIR, name)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f'    [OK] 已保存 {path} ({os.path.getsize(path)} B)')


def _load(name):
    path = os.path.join(DATA_DIR, name)
    if not os.path.exists(path):
        return None
    with open(path, encoding='utf-8') as f:
        return json.load(f)


def _cache_fresh(name, max_age_hours=12):
    path = os.path.join(DATA_DIR, name)
    if not os.path.exists(path):
        return False
    age = time.time() - os.path.getmtime(path)
    return age < max_age_hours * 3600


# ============ 1. LMArena (arena.ai) ============

def fetch_lmarena(force=False):
    """抓取 LMArena 官方 HF 排行榜数据集 (Elo/rating + rank)"""
    print('\n[1] LMArena (arena.ai) Chatbot Arena 排行榜...')
    out_name = 'lmarena.json'
    if not force and _cache_fresh(out_name):
        print('    [SKIP] 缓存未过期 (12h)')
        return _load(out_name)

    configs = {
        'overall': 'text',
        'vision': 'vision',
        'image': 'text_to_image',
        'agent': 'agent',
    }
    # 主数据: text 分类（含 overall + 多语言 + 行业等 29 分类）
    try:
        import pyarrow.parquet as pq
    except ImportError:
        print('    [WARN] 缺少 pyarrow，尝试 pip install pyarrow')
        return None

    url = ('https://huggingface.co/datasets/lmarena-ai/leaderboard-dataset/'
           'resolve/main/text/latest-00000-of-00001.parquet')
    print(f'    下载: {url}')
    raw = _fetch(url, timeout=60)
    if not raw:
        print('    [ERROR] 下载失败')
        return None

    tmp = os.path.join(DATA_DIR, '_lmarena_tmp.parquet')
    with open(tmp, 'wb') as f:
        f.write(raw)
    table = pq.read_table(tmp)
    df = table.to_pandas()
    os.remove(tmp)

    # 整理成按分类分组的榜单
    boards = {}
    for cat, grp in df.groupby('category'):
        rows = []
        for _, r in grp.iterrows():
            rows.append({
                'model': str(r['model_name']),
                'org': str(r['organization']),
                'license': str(r['license']),
                'rating': round(float(r['rating']), 1),
                'rating_ci': [round(float(r['rating_lower']), 1),
                              round(float(r['rating_upper']), 1)],
                'vote_count': int(r['vote_count']),
                'rank': int(r['rank']),
            })
        # 按 rank 排序
        rows.sort(key=lambda x: x['rank'])
        boards[cat] = rows

    data = {
        'source': 'LMArena Chatbot Arena (arena.ai)',
        'url': 'https://arena.ai/',
        'metric': 'Elo rating',
        'updated': df['leaderboard_publish_date'].iloc[0] if 'leaderboard_publish_date' in df.columns else '',
        'models_count': len(df),
        'boards': boards,
    }
    _save(out_name, data)
    return data


# ============ 2. LiveBench ============

def fetch_livebench(force=False):
    """抓取 LiveBench 排行榜 (公开 CSV，按类别聚合为 overall + 7 大类)"""
    print('\n[2] LiveBench 综合智能排行榜...')
    out_name = 'livebench.json'
    if not force and _cache_fresh(out_name):
        print('    [SKIP] 缓存未过期 (12h)')
        return _load(out_name)

    # 选最新可用的 release 版本
    chosen = None
    for rel in reversed(LIVEBENCH_RELEASES):
        url = f'https://livebench.ai/table_{rel}.csv'
        raw = _fetch(url, timeout=15)
        if raw:
            chosen = rel
            break
    if not chosen:
        print('    [ERROR] 未找到可用的 LiveBench 版本')
        return None
    print(f'    使用版本: {chosen}')

    table_raw = _fetch(f'https://livebench.ai/table_{chosen}.csv')
    cat_raw = _fetch(f'https://livebench.ai/categories_{chosen}.json')
    if not table_raw or not cat_raw:
        print('    [ERROR] 数据文件下载失败')
        return None

    table_text = table_raw.decode('utf-8', errors='ignore')
    categories = json.loads(cat_raw.decode('utf-8', errors='ignore'))

    # 解析 CSV
    lines = [l for l in table_text.strip().splitlines() if l.strip()]
    header = [h.strip() for h in lines[0].split(',')]
    rows = []
    for line in lines[1:]:
        # 简单 CSV 解析（LiveBench 字段无逗号）
        cells = [c.strip() for c in line.split(',')]
        if len(cells) == len(header):
            rows.append(dict(zip(header, cells)))

    # 分类聚合: subtask -> 大类
    cat_map = {}
    for big, subs in categories.items():
        for s in subs:
            cat_map[s] = big

    models = []
    for r in rows:
        model = r.get('model', '')
        if not model:
            continue
        entry = {'model': model}
        for big in categories.keys():
            subs = categories[big]
            vals = []
            for s in subs:
                v = r.get(s)
                if v and v != 'None' and re.match(r'^-?\d+\.?\d*$', v):
                    vals.append(float(v))
            entry[big] = round(sum(vals) / len(vals), 1) if vals else None
        # overall = 所有子任务均值
        all_vals = []
        for s in header[1:]:
            v = r.get(s)
            if v and v != 'None' and re.match(r'^-?\d+\.?\d*$', v):
                all_vals.append(float(v))
        entry['Overall'] = round(sum(all_vals) / len(all_vals), 1) if all_vals else None
        models.append(entry)

    # 按 Overall 排序
    models.sort(key=lambda x: x.get('Overall') or 0, reverse=True)

    data = {
        'source': 'LiveBench',
        'url': 'https://livebench.ai/',
        'metric': 'score (平均准确率 %)',
        'release': chosen,
        'models_count': len(models),
        'categories': list(categories.keys()) + ['Overall'],
        'models': models,
    }
    _save(out_name, data)
    return data


# ============ 3. SWE-bench ============

def fetch_swebench(force=False):
    """抓取 SWE-bench 代码能力排行榜 (HTML 内嵌 JSON)"""
    print('\n[3] SWE-bench 代码能力排行榜...')
    out_name = 'swebench.json'
    if not force and _cache_fresh(out_name):
        print('    [SKIP] 缓存未过期 (12h)')
        return _load(out_name)

    raw = _fetch('https://swebench.com/')
    if not raw:
        return None
    html = raw.decode('utf-8', errors='ignore')

    scripts = re.findall(r'<script[^>]*type="application/json"[^>]*>(.*?)</script>',
                         html, re.DOTALL)
    data_json = None
    for s in scripts:
        try:
            d = json.loads(s)
            if isinstance(d, list):
                data_json = d
                break
        except Exception:
            continue
    if data_json is None:
        print('    [ERROR] 未找到 SWE-bench 内嵌数据')
        return None

    # data_json: [{name: "bash-only", results: [...]}, ...]
    boards = {}
    for cat in data_json:
        results = cat.get('results', [])
        rows = []
        for r in results:
            rows.append({
                'model': r.get('name', ''),
                'model_display': r.get('model_display', ''),
                'org': r.get('model_org', ''),
                'agent': r.get('agent', ''),
                'date': r.get('date', ''),
                'cost': r.get('cost'),
                'os_model': r.get('os_model'),
                'os_system': r.get('os_system'),
            })
        boards[cat.get('name', '')] = rows

    data = {
        'source': 'SWE-bench',
        'url': 'https://swebench.com/',
        'metric': 'model + agent + cost (resolved 明细在 per_instance_details)',
        'models_count': sum(len(v) for v in boards.values()),
        'boards': boards,
    }
    _save(out_name, data)
    return data


# ============ 4. EQ-Bench ============

def fetch_eqbench(force=False):
    """抓取 EQ-Bench 情感智能排行榜 (JS 数据文件)"""
    print('\n[4] EQ-Bench 情感智能排行榜...')
    out_name = 'eqbench.json'
    if not force and _cache_fresh(out_name):
        print('    [SKIP] 缓存未过期 (12h)')
        return _load(out_name)

    raw = _fetch('https://eqbench.com/eqbench4/eqbench4_data.js')
    if not raw:
        return None
    text = raw.decode('utf-8', errors='ignore')

    # 格式: const EQBENCH4_DATA = {...};
    m = re.search(r'const\s+EQBENCH4_DATA\s*=\s*({.*?});\s*$', text, re.DOTALL)
    if not m:
        print('    [ERROR] 未找到 EQBENCH4_DATA 结构')
        return None
    try:
        data = json.loads(m.group(1))
    except Exception as e:
        print(f'    [ERROR] JSON 解析失败: {e}')
        return None

    # 整理模型 Elo 榜
    models = data.get('models', [])
    rows = []
    for r in models:
        rows.append({
            'model': r.get('model', ''),
            'display': r.get('display', ''),
            'elo': r.get('elo'),
            'ci': [r.get('ci_low'), r.get('ci_high')],
            'n_scenarios': r.get('n_scenarios'),
        })
    rows.sort(key=lambda x: x.get('elo') or 0, reverse=True)

    result = {
        'source': 'EQ-Bench (v4)',
        'url': 'https://eqbench.com/',
        'metric': 'Elo (情感智能)',
        'generated_at': data.get('generated_at', ''),
        'dimensions': [d.get('key') for d in data.get('dimensions', [])],
        'models_count': len(rows),
        'models': rows,
    }
    _save(out_name, result)
    return result


# ============ 5. Artificial Analysis ============

def fetch_artificialanalysis(force=False):
    """抓取 Artificial Analysis AI 指数排行榜
    优先: 官方 API (需在 auth.json 配 x-api-key, 免费100次/天)
    备选: 首页 HTML 内嵌 JSON-LD (Intelligence Index)"""
    print('\n[5] Artificial Analysis AI 指数排行榜...')
    out_name = 'artificialanalysis.json'
    if not force and _cache_fresh(out_name):
        print('    [SKIP] 缓存未过期 (12h)')
        return _load(out_name)

    # 优先官方 API
    api_key = None
    if os.path.exists(AUTH_FILE):
        try:
            with open(AUTH_FILE, encoding='utf-8') as f:
                api_key = json.load(f).get('artificialanalysis_api_key')
        except Exception:
            pass

    if api_key:
        raw = _fetch('https://artificialanalysis.ai/api/v2/language/models/free',
                     headers_extra={'x-api-key': api_key})
        if raw:
            try:
                api = json.loads(raw.decode('utf-8'))
                rows = []
                for m in api.get('data', []):
                    ev = m.get('evaluations', {})
                    rows.append({
                        'model': m.get('name', ''),
                        'slug': m.get('slug', ''),
                        'org': (m.get('model_creator') or {}).get('name', ''),
                        'intelligence_index': ev.get('artificial_analysis_intelligence_index'),
                        'coding_index': ev.get('artificial_analysis_coding_index'),
                        'agentic_index': ev.get('artificial_analysis_agentic_index'),
                        'input_price': (m.get('pricing') or {}).get('price_1m_input_tokens'),
                        'output_price': (m.get('pricing') or {}).get('price_1m_output_tokens'),
                    })
                rows.sort(key=lambda x: x.get('intelligence_index') or 0, reverse=True)
                data = {
                    'source': 'Artificial Analysis (Data API)',
                    'url': 'https://artificialanalysis.ai/',
                    'metric': 'Intelligence Index v4',
                    'models_count': len(rows),
                    'models': rows,
                }
                _save(out_name, data)
                return data
            except Exception as e:
                print(f'    [WARN] API 解析失败，改用 HTML: {e}')

    # 备选: 首页 HTML 内嵌 JSON-LD
    print('    (无 API key 或 API 失败，使用首页 HTML 内嵌数据)')
    raw = _fetch('https://artificialanalysis.ai/')
    if not raw:
        return None
    html = raw.decode('utf-8', errors='ignore')

    # 提取 JSON-LD 数据（intelligence index 数组）
    # 模式: {"label":"...","artificialAnalysisIntelligenceIndex":xx,"detailsUrl":"..."}
    pattern = re.compile(
        r'\\"label\\":\\"([^\\"]+)\\"'
        r'\\,\\"artificialAnalysisIntelligenceIndex\\":([\d.]+)'
        r'\\,\\"detailsUrl\\":\\"([^\\"]+)\\"')
    matches = pattern.findall(html)
    if not matches:
        # 尝试未转义版本
        pattern2 = re.compile(
            r'"label":"([^"]+)","artificialAnalysisIntelligenceIndex":([\d.]+),"detailsUrl":"([^"]+)"')
        matches = pattern2.findall(html)

    if not matches:
        print('    [ERROR] 未找到 Intelligence Index 数据')
        return None

    seen = {}
    for label, score, url in matches:
        if label not in seen:
            seen[label] = {'model': label, 'intelligence_index': float(score),
                           'url': url}
    rows = sorted(seen.values(), key=lambda x: x['intelligence_index'], reverse=True)

    data = {
        'source': 'Artificial Analysis (HTML 内嵌 Intelligence Index)',
        'url': 'https://artificialanalysis.ai/',
        'metric': 'Intelligence Index (越高越好)',
        'models_count': len(rows),
        'models': rows,
    }
    _save(out_name, data)
    return data


# ============ 6. OpenCompass 司南 ============

def fetch_opencompass(force=False):
    """抓取 OpenCompass 司南排行榜（上海AI实验室）
    数据: OSS 公开 JSON，7 个能力表 (Overall/Language/Knowledge/Reason/Math/Code/Agent)"""
    print('\n[6] OpenCompass 司南 (上海AI实验室)...')
    out_name = 'opencompass.json'
    if not force and _cache_fresh(out_name):
        print('    [SKIP] 缓存未过期 (12h)')
        return _load(out_name)

    raw = _fetch('https://opencompass.oss-cn-shanghai.aliyuncs.com/assets/llm/'
                 'data-llm-ability_official.json')
    if not raw:
        return None
    try:
        data = json.loads(raw.decode('utf-8'))
    except Exception as e:
        print(f'    [ERROR] JSON 解析失败: {e}')
        return None

    # 7 个能力表
    table_keys = {
        'Overall': 'OverallTable', 'Language': 'LanguageTable',
        'Knowledge': 'KnowledgeTable', 'Reasoning': 'ReasonTable',
        'Math': 'MathTable', 'Coding': 'CodeTable', 'Agentic': 'AgentTable',
    }
    boards = {}
    for cat, key in table_keys.items():
        rows = []
        for r in data.get(key, []):
            rows.append({
                'model': r.get('model', ''),
                'org': r.get('org', ''),
                'open_source': r.get('open_source', ''),
                'para_num': r.get('para_num', ''),
                'date': r.get('date', ''),
                'update_date': r.get('update_date', ''),
                'chat_or_base': r.get('chat_or_base', ''),
                'score': r.get('Average'),
                'rank_change': r.get('rank_change', ''),
            })
        rows.sort(key=lambda x: x.get('score') or 0, reverse=True)
        for i, r in enumerate(rows, 1):
            r['rank'] = i
        boards[cat] = rows

    result = {
        'source': 'OpenCompass 司南 (上海AI实验室)',
        'url': 'https://rank.opencompass.org.cn/',
        'metric': 'Average 综合分 (0-100)',
        'updated': data.get('update_date', ''),
        'models_count': len(boards.get('Overall', [])),
        'boards': boards,
    }
    _save(out_name, result)
    return result


# ============ 7. SuperCLUE ============

def fetch_superclue(force=False):
    """抓取 SuperCLUE 中文大模型排行榜
    数据: 官网公开 XLSX（总榜/推理模型榜/推理任务榜/开源榜）"""
    print('\n[7] SuperCLUE 中文大模型排行榜...')
    out_name = 'superclue.json'
    if not force and _cache_fresh(out_name):
        print('    [SKIP] 缓存未过期 (12h)')
        return _load(out_name)

    try:
        import openpyxl
    except ImportError:
        print('    [WARN] 缺少 openpyxl，运行 pip install openpyxl')
        return None

    # 探测最新月份 XLSX（当前月 + 往前推 6 个月）
    from datetime import date, timedelta
    found = None
    for delta in range(0, 7):
        d = date.today() - timedelta(days=delta * 32)
        month_str = f'{d.year}年{d.month}月'
        url = f'https://www.superclueai.com/data/generalboard/' \
              f'{urllib.request.quote(month_str)}.xlsx'
        raw = _fetch(url, timeout=20)
        if raw and len(raw) > 1000 and raw[:2] == b'PK':
            found = (month_str, raw)
            break
    if not found:
        print('    [ERROR] 未找到可用的 SuperCLUE XLSX（尝试 2026年8月 等最新月份）')
        return None
    month_str, raw = found
    print(f'    使用月份: {month_str} ({len(raw)} B)')

    import io
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    boards = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = []
        headers = None
        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = [str(h).strip() for h in row if h is not None]
                continue
            vals = [v for v in row if v is not None]
            if not vals:
                continue
            entry = {}
            for i, h in enumerate(headers):
                entry[h] = row[i] if i < len(row) else None
            rows.append(entry)
        boards[sheet] = rows

    result = {
        'source': 'SuperCLUE',
        'url': 'https://www.superclueai.com/',
        'metric': '总分 (0-100，含数学推理/幻觉控制/科学推理/智能体等维度)',
        'month': month_str,
        'models_count': len(boards.get('总排行榜', [])),
        'boards': boards,
    }
    _save(out_name, result)
    return result


# ============ 8. MMLU-Pro ============

def fetch_mmlupro(force=False):
    """抓取 MMLU-Pro 排行榜（UIUC TIGER-Lab）
    数据: HF Space gradio config（262 模型 × 14 学科）"""
    print('\n[8] MMLU-Pro (UIUC TIGER-Lab)...')
    out_name = 'mmlupro.json'
    if not force and _cache_fresh(out_name):
        print('    [SKIP] 缓存未过期 (12h)')
        return _load(out_name)

    raw = _fetch('https://tiger-lab-mmlu-pro.hf.space/config')
    if not raw:
        # 备用域名
        raw = _fetch('https://tiger-lab-mmlu-pro.hf.space/config')
    if not raw:
        return None
    try:
        cfg = json.loads(raw.decode('utf-8'))
    except Exception as e:
        print(f'    [ERROR] gradio config 解析失败: {e}')
        return None

    # 找 dataframe 组件
    lb_data = None
    lb_headers = None
    for comp in cfg.get('components', []):
        if comp.get('type') != 'dataframe':
            continue
        props = comp.get('props', {})
        val = props.get('value')
        if isinstance(val, dict) and 'data' in val:
            lb_data = val['data']
            lb_headers = val.get('headers', [])
            break
        if isinstance(val, list):
            lb_data = val
            break
    if not lb_data:
        print('    [ERROR] 未找到 dataframe 组件')
        return None

    rows = []
    for r in lb_data:
        if not isinstance(r, list) or not r:
            continue
        model = r[0]
        entry = {'model': model, 'overall': r[3] if len(r) > 3 else None}
        # 学科列从 index 5 开始
        if lb_headers:
            for i, h in enumerate(lb_headers[5:], start=5):
                if i < len(r):
                    entry[h] = r[i]
        rows.append(entry)
    rows.sort(key=lambda x: float(x.get('overall') or 0), reverse=True)

    result = {
        'source': 'MMLU-Pro (UIUC TIGER-Lab)',
        'url': 'https://huggingface.co/spaces/TIGER-Lab/MMLU-Pro',
        'metric': 'Overall 准确率 (0-1)',
        'models_count': len(rows),
        'subjects': lb_headers[5:] if lb_headers else [],
        'models': rows,
    }
    _save(out_name, result)
    return result


# ============ 通用: evals.report 解析 ============

def _parse_evals_report(url):
    """解析 evals.report benchmark 页面表格
    返回 [{model, lab, score, source_model, status, date}]"""
    raw = _fetch(url)
    if not raw:
        return None
    html = raw.decode('utf-8', errors='ignore')

    # 提取 score-table
    m = re.search(r'<table class="score-table"[^>]*>([\s\S]*?)</table>', html)
    if not m:
        print(f'    [ERROR] 未找到 score-table: {url}')
        return None
    table = m.group(1)

    rows = []
    # 每行 <tr>...</tr>
    for tr in re.findall(r'<tr[^>]*>([\s\S]*?)</tr>', table):
        tds = re.findall(r'<td[^>]*>([\s\S]*?)</td>', tr)
        if len(tds) < 5:
            continue
        # Model（含 link 和 Open badge）
        model_td = tds[0]
        model_m = re.search(r'<a[^>]*>([^<]+)</a>', model_td)
        model = model_m.group(1).strip() if model_m else re.sub(r'<[^>]+>', '', model_td).strip()
        is_open = 'badge--open' in model_td or 'model-tag' in model_td
        # Lab
        lab = re.sub(r'<[^>]+>', '', tds[1]).strip()
        # Score
        score = re.sub(r'<[^>]+>', '', tds[2]).strip()
        # Source model
        source_model = re.sub(r'<[^>]+>', '', tds[3]).strip()
        # Status
        status_td = tds[4]
        status_m = re.search(r'badge[^"]*"[^>]*>([^<]+)</span>', status_td)
        status = status_m.group(1).strip() if status_m else re.sub(r'<[^>]+>', '', status_td).strip()
        # Date
        date = re.sub(r'<[^>]+>', '', tds[5]).strip() if len(tds) > 5 else ''

        # 解析分数（91.9% / 80.0）
        score_val = None
        sm = re.search(r'([\d.]+)\s*%', score)
        if sm:
            score_val = float(sm.group(1))
        else:
            sm2 = re.search(r'([\d.]+)', score)
            if sm2:
                score_val = float(sm2.group(1))

        rows.append({
            'model': model,
            'lab': lab,
            'score': score_val,
            'score_display': score,
            'source_model': source_model,
            'status': status,
            'is_open': is_open,
            'date': date,
        })

    # 按分数降序
    rows.sort(key=lambda x: x.get('score') or 0, reverse=True)
    for i, r in enumerate(rows, 1):
        r['rank'] = i
    return rows


# ============ 9. GPQA Diamond ============

def fetch_gpqa(force=False):
    """抓取 GPQA Diamond 排行榜（博士级科学推理，evals.report 官方验证）"""
    print('\n[9] GPQA Diamond (博士级科学推理)...')
    out_name = 'gpqa.json'
    if not force and _cache_fresh(out_name):
        print('    [SKIP] 缓存未过期 (12h)')
        return _load(out_name)

    rows = _parse_evals_report('https://evals.report/benchmarks/gpqa-diamond')
    if not rows:
        return None
    data = {
        'source': 'GPQA Diamond',
        'url': 'https://evals.report/benchmarks/gpqa-diamond',
        'metric': '准确率 %（博士级多学科问答）',
        'models_count': len(rows),
        'models': rows,
    }
    _save(out_name, data)
    return data


# ============ 10. Terminal-Bench 2.1 ============

def fetch_terminalbench(force=False):
    """抓取 Terminal-Bench 2.1 排行榜（CLI 智能体，evals.report 官方验证）"""
    print('\n[10] Terminal-Bench 2.1 (CLI 智能体)...')
    out_name = 'terminalbench.json'
    if not force and _cache_fresh(out_name):
        print('    [SKIP] 缓存未过期 (12h)')
        return _load(out_name)

    rows = _parse_evals_report('https://evals.report/benchmarks/terminal-bench')
    if not rows:
        return None
    data = {
        'source': 'Terminal-Bench 2.1',
        'url': 'https://evals.report/benchmarks/terminal-bench',
        'metric': '任务成功率 %（终端/命令行智能体）',
        'models_count': len(rows),
        'models': rows,
    }
    _save(out_name, data)
    return data


# ============ 11. SWE-bench Pro ============

def fetch_swebenchpro(force=False):
    """抓取 SWE-bench Pro 排行榜（企业级代码，Scale AI，evals.report 验证）"""
    print('\n[11] SWE-bench Pro (企业级代码)...')
    out_name = 'swebenchpro.json'
    if not force and _cache_fresh(out_name):
        print('    [SKIP] 缓存未过期 (12h)')
        return _load(out_name)

    rows = _parse_evals_report('https://evals.report/benchmarks/swe-bench-pro')
    if not rows:
        return None
    data = {
        'source': 'SWE-bench Pro (Scale AI)',
        'url': 'https://evals.report/benchmarks/swe-bench-pro',
        'metric': 'resolved %（企业级软件工程任务）',
        'models_count': len(rows),
        'models': rows,
    }
    _save(out_name, data)
    return data


# ============ 12. BrowseComp-Plus ============

def fetch_browsecomp(force=False):
    """抓取 BrowseComp-Plus 排行榜（网页浏览智能体）
    数据: HF Space 官方 leaderboard.json（LLM+Retriever 组合，84 行）"""
    print('\n[12] BrowseComp-Plus (网页浏览智能体)...')
    out_name = 'browsecomp.json'
    if not force and _cache_fresh(out_name):
        print('    [SKIP] 缓存未过期 (12h)')
        return _load(out_name)

    raw = _fetch('https://tevatron-browsecomp-plus.hf.space/data/leaderboard.json')
    if not raw:
        return None
    try:
        rows_raw = json.loads(raw.decode('utf-8'))
    except Exception as e:
        print(f'    [ERROR] JSON 解析失败: {e}')
        return None

    rows = []
    for r in rows_raw:
        rows.append({
            'model': r.get('LLM', ''),
            'retriever': r.get('Retriever', ''),
            'submitted_by': r.get('Submitted By', ''),
            'scaffold': r.get('Scaffold', ''),
            'accuracy': r.get('Accuracy (%)'),
            'recall': r.get('Recall (%)'),
            'search_calls': r.get('Search Calls'),
            'calib_error': r.get('Calibration Error (%)'),
            'open_weights': r.get('Open Weights?'),
            'eval_date': r.get('Evaluation Date', ''),
        })
    # 按 accuracy 降序
    rows.sort(key=lambda x: x.get('accuracy') or 0, reverse=True)
    for i, r in enumerate(rows, 1):
        r['rank'] = i

    data = {
        'source': 'BrowseComp-Plus',
        'url': 'https://huggingface.co/spaces/Tevatron/BrowseComp-Plus',
        'metric': 'Accuracy %（网页浏览智能体，LLM+Retriever 组合）',
        'models_count': len(rows),
        'models': rows,
    }
    _save(out_name, data)
    return data


# ============ 入口 ============

FETCHERS = {
    'lmarena': fetch_lmarena,
    'livebench': fetch_livebench,
    'swebench': fetch_swebench,
    'eqbench': fetch_eqbench,
    'artificialanalysis': fetch_artificialanalysis,
    'opencompass': fetch_opencompass,
    'superclue': fetch_superclue,
    'mmlupro': fetch_mmlupro,
    'gpqa': fetch_gpqa,
    'terminalbench': fetch_terminalbench,
    'swebenchpro': fetch_swebenchpro,
    'browsecomp': fetch_browsecomp,
}

if __name__ == '__main__':
    import sys
    force = '--force' in sys.argv
    targets = [a for a in sys.argv[1:] if not a.startswith('--')]

    print('=' * 60)
    print('模型排行榜聚合抓取')
    print('=' * 60)
    results = {}
    if targets:
        for t in targets:
            if t in FETCHERS:
                results[t] = FETCHERS[t](force=force)
    else:
        for name, fn in FETCHERS.items():
            results[name] = fn(force=force)

    print('\n' + '=' * 60)
    print('抓取完成汇总:')
    for name, data in results.items():
        if data:
            cnt = data.get('models_count', '?')
            print(f'  [OK] {name}: {cnt} 个模型')
        else:
            print(f'  [FAIL] {name}: 失败')
    print(f'\n数据目录: {DATA_DIR}')
